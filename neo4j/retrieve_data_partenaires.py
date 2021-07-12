import openpyxl
from Python_class import *
from retrieve_data import *
from retrieve_data_isg import *


def parse_partenaire(projects, sheet):
    row = count_row(sheet)
    col = sheet.max_column
    for i in range(2, row + 1):
        index = 0
        partenaire = Partenaire()
        for j in range(1, col + 1):
            if (j == 1):
                index = get_project_index(projects, sheet.cell(i, j).value)
            if (j == 2):
                partenaire.lastname = sheet.cell(i, j).value
            if (j == 3):
                partenaire.firstname = sheet.cell(i, j).value
            if (j == 4):
                partenaire.job = sheet.cell(i, j).value
            if (j == 5):
                partenaire.email = sheet.cell(i, j).value
            if (j == 6):
                partenaire.number = sheet.cell(i, j).value
        projects[index].partenaire = partenaire
    return projects


def parse_excel_partenaire(projects):
    wb = openpyxl.load_workbook("ressources/2021 PLANETE SOLIDAIRE - CONTACTS PARTENAIRES.xlsx")
    sheet = wb.worksheets[0]
    parse_partenaire(projects, sheet)
    return projects
