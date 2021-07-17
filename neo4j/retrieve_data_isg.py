import openpyxl
from retrieve_data_epita import *


def count_row(sheet):
    """
    Function to count the maximum number of rows in an excel sheet

    :param sheet: excel sheet we need the number of rows
    :return: maximum number of rows
    """
    i = 1
    while sheet.cell(i, 1).value is not None:
        i += 1
    return i - 1


# check if the project is already in the projects list
def test_project(projects, project):
    """
    Function to check if the project is already in our projects list

    :param projects: list of projects
    :param project: project we need to check
    :return: boolean saying whether the project is already in the list or not
    """
    for p in projects:
        if p.project_name == project:
            return True
    return False


# get the index of project in projects list
def get_project_index(projects, project):
    """
    Function to get the index of the project in the projects list

    :param projects: list of all the project
    :param project: project we need to check index of
    :return: index of the project
    """
    for p in projects:
        if p.project_name == project:
            return projects.index(p)


# check if the sherpa was already pushed in the project's sherpas list
def check_sherpa(project, sherpa):
    """
    Function to check if the sherpa is already in the project's sherpas list

    :param project: project we need to check the sherpas list
    :param sherpa: sherpa we want to checl
    :return: boolean to indicate whether the sherpa is in the project list or not
    """
    for s in project.sherpas:
        if s.human.lastname == sherpa.split(" ")[0] and s.human.firstname == sherpa.split(" ")[1]:
            return True
    return False


# function to parse a sheet in an excel containing multiple sheets
def parse_sheet(projects, sheet, sheetname):
    """
    Function to parse one sheet in an excel file containing multiple sheets

    :param projects: list of the projects
    :param sheet: excel sheet we run our code on
    :param sheetname: name of the sheet we run the code on
    :return: the list of the projects newly updated with previously missing ones
    """
    row = count_row(sheet)
    col = sheet.max_column
    for i in range(2, row):
        current_project = ''
        index = 0
        sherpa1 = Sherpa()
        sherpa2 = Sherpa()
        student = Pioupiou()
        for j in range(2, col + 1):
            if j == 2:
                student.human.lastname = sheet.cell(i, j).value
            if j == 3:
                student.human.firstname = sheet.cell(i, j).value
            if j == 4:
                student.human.email = sheet.cell(i, j).value
            if j == 5:
                student.team = sheet.cell(i, j).value
            if j == 6:
                if not test_project(projects, sheet.cell(i, j).value):
                    project = Project()
                    project.project_name = sheet.cell(i, j).value
                    projects.append(project)
                index = get_project_index(projects, sheet.cell(i, j).value)
                student.campus = sheetname
                projects[index].students.append(student)
                if sheet.cell(i, j).value != current_project:
                    current_project = sheet.cell(i, j).value
                index = get_project_index(projects, current_project)
            if j == 7:
                sherpa1.human.lastname = sheet.cell(i, j).value.split(" ")[0]
                sherpa1.human.firstname = sheet.cell(i, j).value.split(" ")[1]
                if not check_sherpa(projects[index], sheet.cell(i, j).value):
                    sherpa1.campus = sheetname
                    projects[index].sherpas.append(sherpa1)
            if j == 8:
                sherpa2.human.lastname = sheet.cell(i, j).value.split(" ")[0]
                sherpa2.human.firstname = sheet.cell(i, j).value.split(" ")[1]
                if not check_sherpa(projects[index], sheet.cell(i, j).value):
                    sherpa2.campus = sheetname
                    projects[index].sherpas.append(sherpa2)
            if j == 9:
                student.mission = sheet.cell(i, j).value
                sherpa1.mission = sheet.cell(i, j).value
                sherpa2.mission = sheet.cell(i, j).value
    return projects


# quick print for debuging
def print_projects(projects):
    """
    Function to print every project from the projects list with its sherpas and students

    :param projects: list of all the projects
    :return: NONE
    """
    for project in projects:
        print("Nom du projet : ", project.project_name)
        print("Nombre de sherpa : ", len(project.sherpas))
        for sherpa in project.sherpas:
            print("Sherpa : ", sherpa.human.firstname)
        for student in project.students:
            print("Etudiant : ", student.human.firstname)
        print('\n')


# function to parse all sheets in the isg excel
def parse_excel_isg(projects):
    """
    Function to parse all sheets in the given excel

    :param projects: list of all the projects
    :return: list of updated projects we got from parse_sheet function
    """
    wb = openpyxl.load_workbook("ressources/effectif_campus_clean.xlsx")
    for n in range(0, 6):
        sheet = wb.worksheets[n]
        sheetname = wb.sheetnames[n].split('_')[1]
        parse_sheet(projects, sheet, sheetname)
    return projects
